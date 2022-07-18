from email import header
import requests
from bs4 import BeautifulSoup
import pandas as pd
from time import sleep
import datetime
import os
import smtplib
import email
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
import time
from openpyxl import load_workbook
import helpers






def send_email(email_address: str, body: str):
    """
    send_email sends an email to the email address specified in the
    argument.
    Parameters
    ----------
    email_address: email address of the recipient
    subject: subject of the email
    body: body of the email
    """
    
    msg = MIMEMultipart("alternative")
    msg["Subject"] = "multipart test"
    msg["From"] = "Datatrooper"
    msg["To"] = 'tomasechavarriab@gmail.com'
    msg.attach(MIMEText(body, "html"))


    my_mail = "Datatroopermailservice@gmail.com"
    mypswd = 'datatroopermail336'
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(my_mail, mypswd)
    server.sendmail(my_mail, email_address, msg.as_string())
    server.quit()





while 1:
    #try:
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:94.0) Gecko/20100101 Firefox/94.0',
        'Accept-Language': 'en-US, en;q=0.5'
    }

    df_in = pd.read_excel('input.xlsx')

    query_product = df_in["Producto"].values[0]

    base_url = f'https://www.amazon.com/s?k={query_product}'
    response = requests.get(base_url, headers=headers)
    soup = BeautifulSoup(response.content, 'html.parser')
    results = soup.find_all('a', {'class': 'a-size-base a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal'})

    df = pd.DataFrame(columns=['Fecha', 'Busqueda', 'Producto', 'Precio', 'Descuento', 'Link'])

    for res in results:
        try:
            res_disc = res.find('span', {'class': 'a-size-extra-large s-color-discount s-light-weight-text'}).text
            link = f'https://www.amazon.com/-/es{res["href"]}'
            price = res.find('span', {'class': 'a-offscreen'}).text
            name = res.parent.parent.parent.find('span', {'class': 'a-size-base-plus a-color-base a-text-normal'}).text

            df = df.append({'Fecha': datetime.datetime.now(), 'Busqueda': query_product, 'Producto': name, 'Precio': price, 'Descuento': res_disc, 'Link': link}, ignore_index=True)
        except Exception as e:
            pass

    df['Descuento'] = df['Descuento'].str.replace('%', '').str.replace('-', '').astype(float)
    df.sort_values(by=['Descuento'], inplace=True, ascending=False)

    umbral = df_in['Umbral'].values[0]

    df = df[df['Descuento'] > umbral]

    #writer = pd.ExcelWriter('Precios amazon.xlsx', engine='openpyxl')
    #book = load_workbook('Precios amazon.xlsx') # esto genera error
     # mirar esto: https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas/38075046#38075046
    
    #writer.book = book
    #print(writer.sheets['Sheet1'])

    df.reset_index(drop=True, inplace=True)
    if os.path.isfile('Precios amazon.xlsx'):
        helpers.append_df_to_excel('Precios amazon.xlsx', df, header=False, index=False)
    else:
        helpers.append_df_to_excel('Precios amazon.xlsx', df, index=False)
    #try:
    #    df.to_excel('Precios amazon.xlsx', index=False, startrow=writer.sheets['Sheet1'].max_row)
    #except:
    #    df.to_excel('Precios amazon.xlsx', index=False)

    
    print('Resultados guardados en Precios amazon.xlsx')

    break
    print('Esperando para volver a buscar...')
    time.sleep(60)
    #except Exception as e:
        
    #    print('Error: ', e)
    #    time.sleep(60*10)

